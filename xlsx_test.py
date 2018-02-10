def test_export(num,header,new_value):
    from openpyxl import load_workbook
    wb = load_workbook(filename = 'Exploratory-Student-Info.xlsx')
    sheet_ranges = wb['Sheet1']
    ws = wb.active
    i = 1
    new_record = True
    header_column = -1
    #if header == 'FIRST NAME':
    #    header_column = 'B'
    
    def filter_nonprintable(text):
        import string
        # Get the difference of all ASCII characters from the set of printable characters
        nonprintable = set([chr(i) for i in range(128)]).difference(string.printable)
        # Use translate to remove all non-printable characters
        return text.translate({ord(character):None for character in nonprintable})
    counter = 0
    from string import ascii_lowercase
    for c in ascii_lowercase:
        if counter == 15:
            break
        counter += 1
        
        if filter_nonprintable(ws[c + '1'].value) == filter_nonprintable(header):
            header_column = c

    while ws['M'+str(i)].value != ws['P1'].value:
        if ws['M'+str(i)].value == phone_num:
            break
        i += 1
    if new_record == True:
        ws['M'+str(i)] = phone_num

    ws[header_column+str(i)] = new_value

    wb.save("Exploratory-Student-Info.xlsx")


if __name__ == '__main__':
    phone_num = 5089039394
    header = 'FIRST NAME'
    new_value = 'Jeremy'
    test_export(phone_num,header,new_value)
    
