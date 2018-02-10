#!/usr/bin/env python

from email.mime.text import MIMEText
import imaplib
import email
import smtplib
import pandas as pd
import math
from openpyxl import load_workbook
from string import ascii_lowercase
import unicodedata, re

def search_and_enter(num,header,new_value):
    wb = load_workbook(filename = 'filename.xlsx')
    sheet_ranges = wb['Sheet1']
    ws = wb.active
    i = 1
    new_record = True
    header_column = -1
    
    counter = 0
    master_headers = ['Club #','FIRST NAME','MIDDLE NAME','LAST NAME','SCHOOL','CLASS','DATE OF BIRTH','RETURNING/NEW','FIRST NAME OF GUARDIAN','MIDDLE NAMEOF GUARDIAN','LAST NAME GUARDIAN','ADDRESS','PHONE NUMBER','OTHER PHONE NUMBER','IMAGE RELEASE']

    for c in ascii_lowercase:
        if counter == 15:
            break
        if master_headers[counter] == header:
            header_column = c
        counter += 1
    while ws['M'+str(i)].value != ws['P1'].value:
        if ws['M'+str(i)].value == num:
            break
        i += 1
    if new_record == True:
        ws['M'+str(i)] = num
    
    ws[header_column+str(i)] = new_value

    wb.save("filename.xlsx")

def read_emails(M):
  res_code, data = M.search(None, "ALL")
  if res_code != 'OK':
      print ("No messages found!")
      return

  for num in data[0].split():
      res_code, data = M.fetch(num, '(RFC822)')
      if res_code != 'OK':
          print ("ERROR getting message", num)
          return
          
      msg = email.message_from_string(data[0][1])
      title = msg['Subject']
      phone_num =int(filter(str.isdigit, title))
      email_to_write_to = email.utils.parseaddr(msg['From'])[1]
      body = ""

      if msg.is_multipart():
          for part in msg.walk():
              ctype = part.get_content_type()
              if ctype == 'text/plain':
                  body = part.get_payload(decode=True)  # decode
                  break
      else:
                  body = msg.get_payload(decode=True)
      key = body.splitlines()[2].split(":")[0]
      value = body.splitlines()[2].split(":")[1]
      search_and_enter(phone_num,key,value)

#wrap this in a loop and execute it once in a while to get info
M = imaplib.IMAP4_SSL('imap.gmail.com')

try:
    #very bad idea to hardcode password 
    #use this instead getpass.getpass()
    #but it saves time
    M.login('ghana.sms.project@gmail.com', 'ghanasmsproject') 

except imaplib.IMAP4.error:
    print ("LOGIN FAILED!!! ")
    # ... exit or deal with failure...

#see what we've got from the mail server'
res_code, mailboxes = M.list()
res_code, data = M.select("INBOX")
if res_code == 'OK':
    read_emails(M)
    M.close()

M.logout()




